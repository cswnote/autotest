import pyautogui as pag
import pyperclip as clip
import openpyxl
import time
import TEK_SCOPE
from datetime import datetime
import pyvisa
import sys
import numpy as np

class KMON():
    def __init__(self, start_file_num, info_file_num = 0, capture_path='d:/downloads/'):
        super().__init__()

        self.limit_x = 0
        self.limit_y = 0

        self.line_space = 19
        self.tx_col_space = 0

        self.sw_size = 0
        self.sw_size_x = 0
        self.sw_size_y = 0
        self.sw_name_x = 0
        self.sw_name_y = 0
        self.sw_x = 0
        self.sw_y = 0
        self.sw_index = 0

        self.pb_size = 0
        self.pb_size_x = 0
        self.pb_size_y = 0
        self.pb_name_x = 0
        self.pb_name_y = 0
        self.pb_x = 0
        self.pb_y = 0
        self.pb_index = 0

        self.packet_size = 0
        self.packet_size_x = 0
        self.packet_size_y = 0
        self.packet_name_x = 0
        self.packet_name_y = 0
        self.packet_value_x = 0
        self.packet_value_y = 0

        self.sw = {}
        self.pb = {}
        self.packets = {}

        self.scope_shot = 0

        # # ====================================================== # #

        self.rm = 0

        self.scope = 0
        self.scope_mode = ''

        self.image_format = 'png'
        self.waveform_format = 'csv'

        self.system_path = 'E:/'

        self.csv_save_delay = 0
        self.previous_time = 0

        self.file_num = int(start_file_num)
        self.save_test_info = {'filename': []}

        # # ====================================================== # #
        self.capture_path = capture_path
        self.capture_x1 = 1250
        self.capture_y1 = 72
        self.capture_x2 = 1500
        self.capture_y2 = 772
        self.capture_width = self.capture_x2 - self.capture_x1
        self.capture_height = self.capture_y2 - self.capture_y1

    def set_default_scope(self, serial_num):
        self.rm = pyvisa.ResourceManager()
        instrument_list = self.rm.list_resources()

        for inst in instrument_list:
            if inst.split('::')[3] == serial_num:
                self.scope = self.rm.open_resource(inst)
                break

        print('information of oscilloscope: ', end=' ')
        try:
            print(self.scope.query('*IDN?'))
        except:
            print("can't find Scope(serial number:", serial_num, ')')
            sys.exit()

        horizontal_info = self.scope.query('HORizontal?')
        record_length = horizontal_info.split(';')[3]

        if record_length == '1000':
            self.csv_save_delay = 2
        elif record_length == '10000':
            self.csv_save_delay = 2.4
        elif record_length == '100000':
            self.csv_save_delay = 3.5
        elif record_length == '1000000':
            self.csv_save_delay = 12        # 이하 검증 안됨
        elif record_length == '5000000':
            self.csv_save_delay = 54
        elif record_length == '10000000':
            self.csv_save_delay = 105
        else:
            print('scope error detect')
            sys.exit()


    def save_file_format(self, **kwargs):
        self.image_format = kwargs.get('image_type', 'PNG')
        image_ink_save = kwargs.get('ink_save', 'ON')
        self.waveform_format = kwargs.get('wave_type', 'SPREADSheet')

        if self.waveform_format == 'csv' or self.waveform_format == 'xlsx' or self.waveform_format == 'xls' or \
                self.waveform_format == 'spreadsheet' or self.waveform_format == 'SPREADSheet':
            self.waveform_format = 'SPREADSheet'

        self.image_format = self.image_format.upper()

        now = datetime.now()
        while(1):
            self.scope.write('SAVe:IMAGe:FILEFormat ' + self.image_format)
            if self.scope.query('SAVe:IMAGe:FILEFormat?') == self.image_format + '\n':
                print("image file format is " + self.image_format + '.')
                break
            if (datetime.now() - now).seconds > 10:
                print('error: scope is not answered or check image format')
                sys.exit()

        now = datetime.now()
        while(1):
            self.scope.write('SAVe:WAVEform:FILEFormat ' + self.waveform_format)
            if self.scope.query('SAVe:WAVEform:FILEFormat?') == 'SPREADSHEET\n':
                print("waveform file format is 'csv'")
                break
            if (datetime.now() - now).seconds > 10:
                print('error: scope is not answered')
                sys.exit()

        self.image_format = self.image_format.lower()
        if self.waveform_format == 'SPREADSheet':
            self.waveform_format = 'csv'

    # 나중에 deco로 kmon capture 구현 하자
    def save_scope_files(self, kmon_capture):
        filename = 'tek' + ('{0:04d}'.format(self.file_num))
        self.save_test_info['filename'].append(filename)
        print(filename)
        self.scope.write("SAVe:IMAGe '" + filename + '.' + self.image_format + "'")
        self.scope.write("SAVe:SETUp '" + filename + ".set'")
        self.scope.write("SAVE:WAVEFORM ALL, '" + filename + '.' + self.waveform_format + "'")

        names = list(self.sw.keys())
        for name in names:
        #     if len(self.sw[name]) == 0 and type(self.sw[name]) != "<class 'int'>":
            self.save_test_info[name].append(self.sw[name])
        #     else:
        #         self.save_test_info[name].append(self.sw[name])

        names = list(self.pb.keys())
        for name in names:
            # if len(self.pb[name]) == 0:
            self.save_test_info[name].append(self.pb[name])
            # else:
            #     self.save_test_info[name].append(self.pb[name])

        names = list(self.packets.keys())
        for name in names:
            # if len(self.packets[name]) == 0:
            #     self.save_test_info[name].append('')
            # else:
            self.save_test_info[name].append(self.packets[name])

        if kmon_capture:
            pag.screenshot(self.capture_path + filename + '_kmonCap' + '.png', region=(self.capture_x1, self.capture_y1,
                                                                          self.capture_width, self.capture_height))

        time.sleep(self.csv_save_delay)

        self.file_num = self.file_num + 1


    def check_filesystem(self):
        self.scope.write("FILESystem:CWD 'E/'")
        print(self.scope.query('FILESystem?'))
        print('make later')


    def scope_run_single_mode(self, mode='single'):
        now = datetime.now()
        if mode == 'continuous':
            while(1):
                self.scope.write('ACQuire:STOPAfter RUNSTop')
                if self.scope.query('ACQuire:STOPAfter?') == 'RUNSTOP\n':
                    self.scope_mode = 'continuous'
                    print("Oscilloscope mode is Continuous mode.")
                    break
                if (datetime.now() - now).seconds > 10:
                    print('error: scope is not answered.')
                    sys.exit()

        if mode == 'single':
            while(1):
                self.scope.write('ACQuire:STOPAfter SEQuence')
                if self.scope.query('ACQuire:STOPAfter?') == 'SEQUENCE\n':
                    self.scope_mode = 'single'
                    print("Oscilloscope mode is Single mode.")
                    break
                if (datetime.now() - now).seconds > 10:
                    print('error: scope is not answered.')
                    sys.exit()


    def scope_on(self):
        self.scope.write('ACQuire:STATE ON')
        if self.scope_mode == 'single':
            time.sleep(2)


    def scope_off(self):
        if self.scope_mode == 'continuous':
            self.scope.write('ACQuire:STATE OFF')
            time.sleep(2)


    def save_test_info_initial(self):
        # self.save_test_info['filename'] = []
        # self.save_test_info['sw'] = []
        # self.save_test_info['pb'] = []

        keys = list(self.sw.keys())
        for i, key in enumerate(keys):
            self.save_test_info[key] = []

        keys = list(self.pb.keys())
        for i, key in enumerate(keys):
            self.save_test_info[key] = []

        keys = list(self.packets.keys())
        for i, key in enumerate(keys):
            self.save_test_info[key] = []


    def save_test_list(self, path, info_file_num):
        wb = openpyxl.Workbook()
        ws = wb.active

        save_test_info_keys = list(self.save_test_info.keys())

        for i in range(1, len(save_test_info_keys) + 1):
            ws.cell(1, i).value = save_test_info_keys[i - 1]

        for i, key in enumerate(save_test_info_keys):
            for j, value in enumerate(self.save_test_info[save_test_info_keys[i]], start=2):
                ws.cell(j, i + 1).value = value
        wb.save(path + 'info_test_' + ('{0:02d}'.format(info_file_num)) + '.xlsx')


    def get_packet_info(self):

        previous_sw_name = 0
        for i in range(self.sw_size):
            pag.moveTo(self.sw_name_x, self.sw_name_y + self.line_space * i)
            pag.click(button='right')
            pag.hotkey('shift', 'end')
            previous_clip = clip.paste()
            pag.hotkey('ctrl', 'c')
            sw_name = clip.paste()
            if sw_name != previous_sw_name and previous_clip != sw_name:
                previous_sw_name = sw_name
                if sw_name == 'item name':
                    self.sw['blank' + str(i)] = 0
                else:
                    self.sw[sw_name] = 0
            else:
                self.sw['blank' + str(i)] = 0

        previous_pb_name = 0
        for i in range(self.pb_size):
            pag.moveTo(self.pb_name_x, self.pb_name_y + self.line_space * i)
            pag.click(button='right')
            pag.hotkey('shift', 'end')
            previous_clip = clip.paste()
            pag.hotkey('ctrl', 'c')
            pb_name = clip.paste()
            if pb_name != previous_pb_name and previous_clip != pb_name:
                previous_pb_name = pb_name
                if pb_name == 'item name':
                    self.pb['blank' + str(i)] = 0
                else:
                    self.pb[pb_name] = 0
            else:
                self.pb['blank' + str(i)] = 0

        previous_packet_name = 0
        for i in range(self.packet_size):
            pag.moveTo(self.packet_name_x, self.packet_name_y + self.line_space * i)
            pag.click(button='right')
            pag.hotkey('shift', 'end')
            previous_clip = clip.paste()
            pag.hotkey('ctrl', 'c')
            packet_name = clip.paste()
            if self.sw_index == i:
                self.sw_index = packet_name
            if self.pb_index == i:
                self.pb_index = packet_name
            if packet_name != previous_packet_name and previous_clip != packet_name:
                previous_packet_name = packet_name
                pag.moveTo(self.packet_name_x + self.tx_col_space, self.packet_name_y + self.line_space * i)
                pag.doubleClick(button='right')
                # previous_clip = clip.paste()
                pag.hotkey('ctrl', 'c')
                self.packets[packet_name] = clip.paste()
            else:
                pag.moveTo(self.packet_name_x + self.tx_col_space, self.packet_name_y + self.line_space * i)
                pag.doubleClick(button='right')
                pag.hotkey('ctrl', 'c')
                self.packets['blank' + str(i)] = clip.paste()

        sw_values = format(int(self.packets[self.sw_index]), 'b')
        sw_keys = list(self.sw.keys())
        for i in range(len(sw_keys) - len(sw_values)):
            sw_values = '0' + sw_values
        sw_values = ''.join(reversed(sw_values))
        for i in range(len(sw_keys)):
            self.sw[sw_keys[i]] = int(sw_values[i])

        pb_values = format(int(self.packets[self.pb_index]), 'b')
        pb_keys = list(self.pb.keys())
        for i in range(len(pb_keys) - len(pb_values)):
            pb_values = '0' + pb_values
        pb_values = ''.join(reversed(pb_values))
        for i in range(len(pb_keys)):
            self.pb[pb_keys[i]] = int(pb_values[i])

        return self.sw, self.pb, self.packets


    def origin_coordinate(self):
        self.limit_x, self.limit_y = pag.size()

        # if self.limit_x == 1920 and self.limit_y == 1080:
        if True:
            self.line_space = 19
            self.tx_col_space = 104

            self.sw_size_x = 565
            self.sw_size_y = 43
            self.sw_name_x = 492
            self.sw_name_y = 82

            self.pb_size_x = 565
            self.pb_name_x = 492

            self.packet_size_x = 703
            self.packet_size_y = 58
            self.packet_name_x = 624
            self.packet_name_y = 82

            pag.moveTo(self.sw_size_x, self.sw_size_y)
            pag.doubleClick(button='right')
            pag.hotkey('ctrl', 'c')
            self.sw_size = int(clip.paste())

            pag.move(0, 17)
            pag.doubleClick(button='right')
            pag.hotkey('ctrl', 'c')
            self.sw_index = int(clip.paste())

            self.pb_size_y = self.sw_name_y + self.line_space * (self.sw_size - 1) + 32
            self.pb_name_y = self.pb_size_y + 42
            pag.moveTo(self.pb_size_x, self.pb_size_y)
            pag.doubleClick(button='right')
            pag.hotkey('ctrl', 'c')
            self.pb_size = int(clip.paste())

            pag.move(0, 17)
            pag.doubleClick(button='right')
            pag.hotkey('ctrl', 'c')
            self.pb_index = int(clip.paste())


            pag.moveTo(self.packet_size_x, self.packet_size_y)
            pag.doubleClick(button='right')
            pag.hotkey('ctrl', 'c')
            self.packet_size = int(clip.paste())

            self.sw_x = self.sw_name_x + self.tx_col_space
            self.sw_y = self.sw_name_y
            self.pb_x = self.pb_name_x + self.tx_col_space
            self.pb_y = self.pb_name_y
            self.packet_value_x = self.packet_name_x + self.tx_col_space
            self.packet_value_y = self.packet_name_y


    def push_button(self, status, name, type):
        if type == 'sw':
            if status != self.sw[name]:
                idx = list(self.sw.keys()).index(name)
                pag.moveTo(self.sw_x, self.sw_y + self.line_space * idx)
                pag.click(button='right')
                self.sw[name] = status
            sw_values = list(self.sw.values())
            sum = 0
            for i in range(len(sw_values)):
                sum = sum + 2**i * sw_values[i]
            self.packets[self.sw_index] = sum

        elif type == 'pb':
            if status != self.pb[name]:
                idx = list(self.pb.keys()).index(name)
                pag.moveTo(self.pb_x, self.pb_y + self.line_space * idx)
                pag.click(button='right')
                self.pb[name] = status
            pb_values = list(self.pb.values())
            sum = 0
            for i in range(len(pb_values)):
                sum = sum + 2**i * pb_values[i]
            self.packets[self.pb_index] = sum


    def change_packet_value(self, value, name):
        idx = list(self.packets.keys()).index(name)
        pag.moveTo(self.packet_value_x, self.packet_value_y + self.line_space * idx)
        pag.doubleClick(button='right')
        pag.write(str(value))
        pag.press('enter')
        self.packets[name] = str(value)


        if name == list(self.packets.keys())[1]:
            set_value = bin(value).split('b')[1]
            a = len(set_value)
            b = len(self.sw)
            if len(set_value) < len(self.sw):
                for i in range(len(self.sw) - len(set_value)):
                    set_value = set_value + '0'
            sw_name = list(self.sw.keys())
            for i, name in enumerate(sw_name):
                self.sw[name] = int(set_value[i])
                # if self.sw[name] != int(set_value[i]):
                    # self.push_button(int(set_value[i]), name, 'sw')

        elif name == list(self.packets.keys())[2]:
            set_value = bin(value).split('b')[1]
            a = len(set_value)
            b = len(self.sw)
            if len(set_value) < len(self.pb):
                for i in range(len(self.pb) - len(set_value)):
                    set_value = set_value + '0'
            pb_name = list(self.pb.keys())
            for i, name in enumerate(pb_name):
                self.pb[name] = int(set_value[i])
                # if self.pb[name] != int(set_value[i]):
                    # self.push_button(int(set_value[i]), name, 'pb')

        print('==============')


    def make_test_sequence(self, file, path):
        wb_test = openpyxl.load_workbook(path + file)
        ws_test = wb_test['test']

        b_length = 0
        c_length = 0

        test_seq = []


        loop_count = 0

        for i in range(1, 10001):
            if ws_test['b' + str(i)].value is None:
                b_length = i - 1
                break

        for i in range(1, 10001):
            if ws_test['c' + str(i)].value is None:
                c_length = i - 1
                break

        if b_length > c_length:
            test_length = b_length
        else:
            test_length = c_length

        for i in range(1, test_length + 1):
            if ws_test.cell(i, 1).value is None:
                if ws_test.cell(i, 2).value.lower() == 'sw':
                    name = ws_test.cell(i, 3).value
                    value = ws_test.cell(i, 4).value
                    test_seq.append(['sw', name, value])
                elif ws_test.cell(i, 2).value.lower() == 'pb':
                    name = ws_test.cell(i, 3).value
                    value = ws_test.cell(i, 4).value
                    test_seq.append(['pb', name, value])
                elif ws_test.cell(i, 2).value.lower() == 'pak':
                    name = ws_test.cell(i, 3).value
                    value = ws_test.cell(i, 4).value
                    test_seq.append(['pak', name, value])
                elif ws_test.cell(i, 2).value.lower() == 'scope':
                    name = ws_test.cell(i, 3).value
                    value = ws_test.cell(i, 4).value
                    test_seq.append(['scope', name, value])
                elif ws_test.cell(i, 2).value.lower() == 'pause':
                    name = ws_test.cell(i, 3).value
                    value = ws_test.cell(i, 4).value
                    test_seq.append(['pause', name, value])

            elif ws_test.cell(i, 1).value.lower() == 'loop on':
                if loop_count == 0:
                    loop_count += 1
                    loop_seq = ['loop_0']
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                        name = ws_test.cell(i, 3).value
                        value_start = ws_test.cell(i, 4).value
                        value_stop = ws_test.cell(i + 1, 4).value
                        value_step = ws_test.cell(i + 2, 4).value
                        loop_seq.append(['pak_loop', name, value_start, value_stop, value_step])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pause', name, value])

                elif loop_count == 1:
                    loop_count += 1
                    loop_seq1 = ['loop_1']
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                        name = ws_test.cell(i, 3).value
                        value_start = ws_test.cell(i, 4).value
                        value_stop = ws_test.cell(i + 1, 4).value
                        value_step = ws_test.cell(i + 2, 4).value
                        loop_seq1.append(['pak_loop', name, value_start, value_stop, value_step])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pause', name, value])

                elif loop_count == 2:
                    loop_count += 1
                    loop_seq2 = ['loop_2']
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                        name = ws_test.cell(i, 3).value
                        value_start = ws_test.cell(i, 4).value
                        value_stop = ws_test.cell(i + 1, 4).value
                        value_step = ws_test.cell(i + 2, 4).value
                        loop_seq2.append(['pak_loop', name, value_start, value_stop, value_step])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pause', name, value])


            elif ws_test.cell(i, 1).value.lower() == '-':
                if ws_test.cell(i, 2).value.lower() == 'pak_stop' or ws_test.cell(i, 2).value.lower() == 'pak_step':
                    pass
                else:
                    if loop_count == 1:
                        if ws_test.cell(i, 2).value.lower() == 'sw':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq.append(['sw', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pb':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq.append(['pb', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq.append(['pak', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                            name = ws_test.cell(i, 3).value
                            value_start = ws_test.cell(i, 4).value
                            value_stop = ws_test.cell(i + 1, 4).value
                            value_step = ws_test.cell(i + 2, 4).value
                            loop_seq.append(['pak_loop', name, value_start, value_stop, value_step])
                        elif ws_test.cell(i, 2).value.lower() == 'scope':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq.append(['scope', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pause':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq.append(['pause', name, value])
                    if loop_count == 2:
                        if ws_test.cell(i, 2).value.lower() == 'sw':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq1.append(['sw', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pb':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq1.append(['pb', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq1.append(['pak', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                            name = ws_test.cell(i, 3).value
                            value_start = ws_test.cell(i, 4).value
                            value_stop = ws_test.cell(i + 1, 4).value
                            value_step = ws_test.cell(i + 2, 4).value
                            loop_seq1.append(['pak_loop', name, value_start, value_stop, value_step])
                        elif ws_test.cell(i, 2).value.lower() == 'scope':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq1.append(['scope', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pause':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq1.append(['pause', name, value])
                    if loop_count == 3:
                        if ws_test.cell(i, 2).value.lower() == 'sw':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq2.append(['sw', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pb':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq2.append(['pb', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq2.append(['pak', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pak_start':
                            name = ws_test.cell(i, 3).value
                            value_start = ws_test.cell(i, 4).value
                            value_stop = ws_test.cell(i + 1, 4).value
                            value_step = ws_test.cell(i + 2, 4).value
                            loop_seq2.append(['pak_loop', name, value_start, value_stop, value_step])
                        elif ws_test.cell(i, 2).value.lower() == 'scope':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq2.append(['scope', name, value])
                        elif ws_test.cell(i, 2).value.lower() == 'pause':
                            name = ws_test.cell(i, 3).value
                            value = ws_test.cell(i, 4).value
                            loop_seq2.append(['pause', name, value])


            elif ws_test.cell(i, 1).value.lower() == 'loop off':
                if loop_count == 1:
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq.append(['pause', name, value])
                    test_seq.append(loop_seq)
                    loop_count -= 1

                elif loop_count == 2:
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq1.append(['pause', name, value])
                    loop_seq.append(loop_seq1)
                    loop_count -= 1

                elif loop_count == 3:
                    if ws_test.cell(i, 2).value.lower() == 'sw':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['sw', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pb':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pb', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pak':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pak', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'scope':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['scope', name, value])
                    elif ws_test.cell(i, 2).value.lower() == 'pause':
                        name = ws_test.cell(i, 3).value
                        value = ws_test.cell(i, 4).value
                        loop_seq2.append(['pause', name, value])
                    loop_seq1.append(loop_seq2)
                    loop_count -= 1
        return test_seq


    def loop_test(self, loop_seq, kmon_capture=False):
        # del loop_seq[0]
        # loop_count = 0
        pak_name = ''
        pak_step = 0
        pak_start = 0
        pak_stop = 0
        iteration = 0
        for idx, seq in enumerate(loop_seq):
            if seq[0] == 'pak_loop':
                pak_name = seq[1]
                iteration = np.arange(seq[2], seq[3], seq[4])
        self.loop_process(pak_name, iteration, loop_seq, kmon_capture)


    def loop_process(self, name, iteration, loop_seq, kmon_capture=False):
        for i in iteration:
            for idx, seq in enumerate(loop_seq):
                if seq[0] == 'sw' or seq[0] == 'pb':
                    self.push_button(seq[2], seq[1], seq[0])
                elif seq[0] == 'pak' and len(seq[0]) == 3:
                    self.change_packet_value(seq[2], seq[1])
                elif seq[0].split('_')[0] == 'pak':
                    self.change_packet_value(round(i, 8), name)
                elif seq[0].split('_')[0] == 'loop':
                    self.loop_test(seq, kmon_capture)
                elif seq[0].lower() == 'scope':
                    if seq[1].lower() == 'on' or seq[1].lower() == 'run':
                        self.scope_on()
                    elif seq[1].lower() == 'off' or seq[1].lower() == 'stop':
                        self.scope_off()
                    elif seq[1].lower() == 'save':
                        self.save_scope_files(kmon_capture)
                elif seq[0] == 'pause':
                    time.sleep(int(seq[2]))

    def test_process(self, test_seq, kmon_capture=False):
        for idx, seq in enumerate(test_seq):
            if seq[0] == 'sw' or seq[0] == 'pb':
                self.push_button(seq[2], seq[1], seq[0])
            elif seq[0] == 'pak':
                self.change_packet_value(seq[2], seq[1])
            elif seq[0].split('_')[0] == 'loop':
                self.loop_test(seq, kmon_capture)
            elif seq[0] == 'scope':
                if seq[1].lower() == 'on' or seq[1].lower() == 'run':
                    self.scope_on()
                elif (seq[1].lower() == 'off' or seq[1].lower() == 'stop') and self.scope_mode == 'continuous':
                    self.scope_off()
                elif seq[1].lower() == 'save':
                    self.save_scope_files(kmon_capture)
            elif seq[0] == 'pause':
                time.sleep(int(seq[2]))
        print('=============================')

    # def save_file_info(self, sheet, file, path):
    #     openpyxl